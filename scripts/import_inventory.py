import argparse
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook


PORT_RE = re.compile(r"\b(6553[0-5]|655[0-2]\d|65[0-4]\d{2}|6[0-4]\d{3}|[1-5]?\d{1,4})\b")
URL_RE = re.compile(r"https?://[^\s;,]+", re.IGNORECASE)
IP_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")


@dataclass
class DeviceRecord:
    hostname: str
    ip_address: str
    role: str = "infrastructure"
    site: str = "home-lab"
    platform: str = ""
    owner: str = "Brandon"
    tags: set[str] = field(default_factory=lambda: {"homelab"})
    connection_type: str = "icmp"
    tcp_ports: set[int] = field(default_factory=set)
    http_urls: set[str] = field(default_factory=set)
    ssh_enabled: bool = False
    notes: list[str] = field(default_factory=list)

    def payload(self) -> dict[str, Any]:
        return {
            "hostname": self.hostname,
            "ip_address": self.ip_address,
            "role": self.role,
            "site": self.site,
            "platform": self.platform or None,
            "owner": self.owner,
            "tags": sorted(self.tags),
            "connection_type": self.connection_type,
            "tcp_ports": sorted(self.tcp_ports),
            "http_urls": sorted(self.http_urls),
            "ssh_enabled": self.ssh_enabled,
            "notes": " | ".join(dict.fromkeys(self.notes))[:2000] or None,
            "is_active": True,
        }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import homelab inventory into NetOps Command Center.")
    parser.add_argument("--workbook", required=True, help="Path to bratsu_network_inventory_updated.xlsx")
    parser.add_argument("--api-base", default="http://127.0.0.1:8000", help="FastAPI base URL")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    records = parse_workbook(workbook_path)
    result = upsert_devices(args.api_base.rstrip("/"), records)
    print(f"Imported inventory from {workbook_path}")
    print(f"Created: {result['created']}")
    print(f"Updated: {result['updated']}")
    print(f"Skipped: {result['skipped']}")
    print(f"Total source devices: {len(records)}")


def parse_workbook(path: Path) -> list[DeviceRecord]:
    wb = load_workbook(path, data_only=True)
    records: dict[str, DeviceRecord] = {}

    def get_record(hostname: str, ip: str, role: str, platform: str = "") -> DeviceRecord:
        ip = clean(ip)
        hostname = slug_hostname(hostname)
        if ip not in records:
            records[ip] = DeviceRecord(hostname=hostname, ip_address=ip, role=role, platform=platform)
        record = records[ip]
        if record.hostname.startswith("device-") and hostname:
            record.hostname = hostname
        if role and record.role == "infrastructure":
            record.role = role
        if platform and not record.platform:
            record.platform = platform
        return record

    if "Host" in wb.sheetnames:
        for row in table_rows(wb["Host"]):
            record = get_record(row["Host"], row["IP"], "proxmox-host")
            record.tags.update({"proxmox", "hypervisor", "critical"})
            add_urls_from_text(record, row.get("Access / Security Notes"))
            record.tcp_ports.update({22, 8006})
            add_note(record, row.get("Description"), row.get("Access / Security Notes"), row.get("Last Verified"))
            record.ssh_enabled = has_ssh(row)

    if "VMs" in wb.sheetnames:
        for row in table_rows(wb["VMs"]):
            record = get_record(row["Name"], row["IP"], "virtual-machine", row.get("OS", ""))
            record.tags.update({"vm", "proxmox"})
            add_urls_from_text(record, row.get("Access / Security Notes"))
            add_note(record, f"VM {row.get('ID')}: {row.get('Description')}", row.get("Access / Security Notes"), row.get("Last Verified"))
            record.ssh_enabled = has_ssh(row)

    if "LXC Containers" in wb.sheetnames:
        for row in table_rows(wb["LXC Containers"]):
            record = get_record(row["Name"], row["IP"], "lxc-container")
            record.tags.update({"lxc", "proxmox"})
            if not is_disabled(row):
                add_ports_from_field(record, row.get("Ports"))
                add_urls_from_text(record, row.get("Access / Security Notes"))
                add_web_urls_for_lxc(record)
            add_note(record, f"CT {row.get('ID')}: {row.get('Description')}", row.get("Access / Security Notes"), row.get("Last Verified"))
            record.ssh_enabled = has_ssh(row)

    if "Other Infrastructure" in wb.sheetnames:
        for row in table_rows(wb["Other Infrastructure"]):
            record = get_record(row["Name"], row["IP"], str(row.get("Type") or "infrastructure"))
            record.tags.update({"infrastructure"})
            add_ports_from_field(record, row.get("Ports / Shares"))
            add_urls_from_text(record, row.get("Access / Security Notes"))
            if "nas" in record.role.lower() or "smb" in record.role.lower():
                record.tcp_ports.add(445)
                record.tags.add("storage")
            if "dns" in record.role.lower():
                record.tags.add("dns")
                record.tcp_ports.discard(53)
                record.tcp_ports.discard(443)
            add_web_urls_for_lxc(record)
            add_note(record, row.get("Description"), row.get("Access / Security Notes"), row.get("Last Verified"))
            record.ssh_enabled = has_ssh(row)

    if "Services" in wb.sheetnames:
        for row in table_rows(wb["Services"]):
            ip = clean(row.get("IP"))
            if not ip or ip not in records:
                host = row.get("Host / CT") or row.get("Service") or f"device-{ip}"
                records[ip] = DeviceRecord(hostname=slug_hostname(str(host).split()[-1]), ip_address=ip)
            record = records[ip]
            service_name = clean(row.get("Service"))
            if service_name:
                record.tags.add(slug_hostname(service_name))
            if not is_disabled(row):
                add_ports_from_field(record, row.get("Port(s)"))
                add_urls_from_text(record, row.get("URL / Endpoint"))
            add_note(record, f"{service_name}: {row.get('Status / Notes')}", None, row.get("Last Verified"))

    apply_known_monitoring_overrides(records)

    return sorted(records.values(), key=lambda item: item.hostname)


def table_rows(sheet) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(header) for header in rows[0]]
    output = []
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers)) if headers[index]}
        if any(value not in (None, "") for value in row.values()):
            output.append(row)
    return output


def upsert_devices(api_base: str, records: list[DeviceRecord]) -> dict[str, int]:
    counts = {"created": 0, "updated": 0, "skipped": 0}
    with httpx.Client(timeout=20.0) as client:
        existing_response = client.get(f"{api_base}/api/v1/devices")
        existing_response.raise_for_status()
        existing = existing_response.json()["items"]
        by_ip = {item["ip_address"]: item for item in existing}

        for record in records:
            payload = record.payload()
            if not IP_RE.match(payload["ip_address"]):
                counts["skipped"] += 1
                continue
            if payload["ip_address"] in by_ip:
                response = client.put(f"{api_base}/api/v1/devices/{by_ip[payload['ip_address']]['id']}", json=payload)
                counts["updated"] += 1
            else:
                response = client.post(f"{api_base}/api/v1/devices", json=payload)
                counts["created"] += 1
            response.raise_for_status()
    return counts


def add_ports_from_field(record: DeviceRecord, value: Any) -> None:
    text = clean(value)
    if not text or text == "-":
        return
    if text.startswith("//"):
        return
    text = text.replace("app ports", "")
    for match in PORT_RE.findall(text):
        port = int(match)
        if 1 <= port <= 65535:
            record.tcp_ports.add(port)


def add_urls_from_text(record: DeviceRecord, value: Any) -> None:
    text = clean(value)
    if not text:
        return
    for match in URL_RE.findall(text):
        url = match.rstrip(".,)")
        record.http_urls.add(url)
        parsed = urlparse(url)
        if parsed.hostname and parsed.hostname != record.ip_address:
            continue
        if parsed.port:
            record.tcp_ports.add(parsed.port)
        elif parsed.scheme == "https":
            record.tcp_ports.add(443)
        elif parsed.scheme == "http":
            record.tcp_ports.add(80)


def add_web_urls_for_lxc(record: DeviceRecord) -> None:
    web_ports = {80, 81, 3000, 5055, 7575, 7777, 7878, 8096, 8265, 8266, 8502, 8989, 9696, 11434}
    for port in record.tcp_ports:
        if port in web_ports:
            record.http_urls.add(f"http://{record.ip_address}:{port}")


def add_note(record: DeviceRecord, description: Any, access_notes: Any, last_verified: Any) -> None:
    parts = [clean(description), clean(access_notes)]
    if last_verified:
        parts.append(f"Last verified: {clean(last_verified)}")
    note = "; ".join(part for part in parts if part)
    if note:
        record.notes.append(note)


def has_ssh(row: dict[str, Any]) -> bool:
    access_notes = clean(row.get("Access / Security Notes")).lower()
    username = clean(row.get("Username"))
    return "ssh" in access_notes or username in {"root", "bratsu"}


def is_disabled(row: dict[str, Any]) -> bool:
    text = " ".join(clean(value).lower() for value in row.values())
    return "disabled" in text or "closed unless intentionally enabled" in text


def apply_known_monitoring_overrides(records: dict[str, DeviceRecord]) -> None:
    proxmox = records.get("192.168.68.72")
    if proxmox:
        proxmox.connection_type = "tcp"
        proxmox.tcp_ports.update({22, 8006})
        proxmox.tags.add("icmp-may-be-blocked")

    npmplus = records.get("192.168.68.63")
    if npmplus:
        npmplus.ssh_enabled = False
        npmplus.notes.append("SSH probe disabled in NetOps because the reverse proxy is monitored by HTTP/TCP service checks.")

    pihole = records.get("192.168.68.50")
    if pihole:
        pihole.tcp_ports.discard(53)
        pihole.tcp_ports.discard(443)
        pihole.tcp_ports.add(80)
        pihole.notes.append("DNS service uses UDP/53; NetOps MVP tracks the Pi-hole web/admin reachability over HTTP/TCP.")

    website = records.get("192.168.68.96")
    if website:
        website.tcp_ports.discard(80)
        website.tcp_ports.discard(443)
        website.notes.append("Public HTTPS URLs are monitored directly; local VM 80/443 may be closed or proxy-routed.")

    mini_pc = records.get("192.168.68.65")
    if mini_pc:
        mini_pc.hostname = "mini-pc"
        mini_pc.role = "mini-pc"
        mini_pc.tags.update({"mini-pc", "physical-host"})
        mini_pc.tcp_ports = {22, 8080}
        mini_pc.http_urls = {"http://192.168.68.65:8080"}
        mini_pc.ssh_enabled = True
        mini_pc.notes.append("NetOps override: live probe found SSH on 22 and HTTP-style service on 8080; workbook HTTP/80 probe is ignored.")


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def slug_hostname(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9-]+", "-", clean(value).lower()).strip("-")
    return cleaned or "device"


if __name__ == "__main__":
    main()
